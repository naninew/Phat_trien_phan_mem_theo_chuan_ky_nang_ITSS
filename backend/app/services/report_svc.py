from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import io

from app.models.user import User
from app.models.company import RescueCompany
from app.models.request import RescueRequest
from app.models.payment import Payment
from app.models.review import Review

_STATUS_BUCKETS = ("PENDING", "IN_PROGRESS", "COMPLETED", "CANCELLED", "REJECTED")
_IN_PROGRESS_STATUSES = frozenset({"ACCEPTED", "ASSIGNED", "ON_THE_WAY", "IN_PROGRESS"})


def _map_request_status(raw_status: str) -> str:
    if raw_status in _IN_PROGRESS_STATUSES:
        return "IN_PROGRESS"
    if raw_status in _STATUS_BUCKETS:
        return raw_status
    return raw_status


def _apply_request_filters(
    query,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    company_id: Optional[int] = None,
    incident_type: Optional[str] = None,
    status: Optional[str] = None,
):
    if from_date:
        query = query.filter(RescueRequest.created_at >= from_date)
    if to_date:
        query = query.filter(RescueRequest.created_at <= to_date)
    if company_id:
        query = query.filter(RescueRequest.company_id == company_id)
    if incident_type:
        query = query.filter(RescueRequest.incident_type == incident_type)
    if status:
        if status == "IN_PROGRESS":
            query = query.filter(RescueRequest.status.in_(_IN_PROGRESS_STATUSES))
        else:
            query = query.filter(RescueRequest.status == status)
    return query


def _apply_payment_date_filters(query, from_date: Optional[datetime], to_date: Optional[datetime]):
    if from_date:
        query = query.filter(Payment.created_at >= from_date)
    if to_date:
        query = query.filter(Payment.created_at <= to_date)
    return query


def _apply_review_date_filters(query, from_date: Optional[datetime], to_date: Optional[datetime]):
    if from_date:
        query = query.filter(Review.created_at >= from_date)
    if to_date:
        query = query.filter(Review.created_at <= to_date)
    return query

def get_admin_stats(db: Session) -> Dict[str, Any]:
    today_str = datetime.utcnow().strftime('%Y-%m-%d')
    month_str = datetime.utcnow().strftime('%Y-%m')

    total_users = db.query(func.count(User.id)).scalar()
    total_companies = db.query(func.count(RescueCompany.id)).scalar()
    active_companies = db.query(func.count(RescueCompany.id)).filter(RescueCompany.status == "active").scalar()
    
    # pending companies
    pending_companies = db.query(func.count(RescueCompany.id)).filter(
        RescueCompany.is_verified == False,
        RescueCompany.status != "suspended"
    ).scalar()

    total_requests = db.query(func.count(RescueRequest.id)).scalar()
    pending_requests = db.query(func.count(RescueRequest.id)).filter(RescueRequest.status == "PENDING").scalar()
    
    # requests today
    requests_today = db.query(func.count(RescueRequest.id)).filter(
        func.strftime('%Y-%m-%d', RescueRequest.created_at) == today_str
    ).scalar()

    # revenue total and this month
    total_revenue = db.query(func.sum(Payment.amount)).filter(Payment.status == "success").scalar() or 0
    revenue_this_month = db.query(func.sum(Payment.amount)).filter(
        Payment.status == "success",
        func.strftime('%Y-%m', Payment.created_at) == month_str
    ).scalar() or 0

    # requests by status
    status_data = db.query(
        RescueRequest.status,
        func.count(RescueRequest.id)
    ).group_by(RescueRequest.status).all()
    requests_by_status = {s[0]: s[1] for s in status_data}
    
    return {
        "total_users": total_users,
        "total_companies": total_companies,
        "active_companies": active_companies,
        "pending_companies": pending_companies,
        "total_requests": total_requests,
        "pending_requests": pending_requests,
        "requests_today": requests_today,
        "total_revenue": float(total_revenue),
        "revenue_this_month": float(revenue_this_month),
        "requests_by_status": requests_by_status
    }

def get_daily_request_stats(db: Session, days: int = 7) -> Dict[str, Any]:
    """Lấy số lượng yêu cầu theo từng ngày trong X ngày gần nhất."""
    from datetime import timedelta
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days - 1)
    
    start_str = start_date.strftime('%Y-%m-%d')
    end_str = end_date.strftime('%Y-%m-%d 23:59:59')

    # Query group by date
    daily_data = db.query(
        func.strftime('%Y-%m-%d', RescueRequest.created_at).label('date'),
        func.count(RescueRequest.id).label('count')
    ).filter(
        RescueRequest.created_at >= start_str,
        RescueRequest.created_at <= end_str
    ).group_by('date').order_by('date').all()

    # Create a full map of the last X days with 0 counts to ensure no gaps
    date_map = {}
    for i in range(days):
        d = (start_date + timedelta(days=i)).strftime('%Y-%m-%d')
        date_map[d] = 0
    
    for r in daily_data:
        date_map[r.date] = r.count

    sorted_dates = sorted(list(date_map.keys()))
    return {
        "labels": sorted_dates,
        "values": [date_map[d] for d in sorted_dates]
    }

def get_chart_stats(db: Session) -> Dict[str, Any]:
    # 1. Doanh thu theo tháng (6 tháng gần nhất)
    # Vì dùng SQLite nên xử lý format ngày hơi khác một chút
    revenue_data = db.query(
        func.strftime('%Y-%m', Payment.created_at).label('month'),
        func.sum(Payment.amount).label('total')
    ).filter(Payment.status == "success").group_by('month').order_by('month').limit(6).all()
    
    revenue_chart = {
        "labels": [r.month for r in revenue_data],
        "values": [float(r.total) for r in revenue_data]
    }

    # 2. Phân bổ trạng thái yêu cầu
    status_data = db.query(
        RescueRequest.status,
        func.count(RescueRequest.id)
    ).group_by(RescueRequest.status).all()
    
    status_chart = [
        {"name": s[0], "value": s[1]} for s in status_data
    ]

    # 3. Loại sự cố phổ biến
    incident_data = db.query(
        RescueRequest.incident_type,
        func.count(RescueRequest.id)
    ).group_by(RescueRequest.incident_type).all()
    
    incident_chart = {
        "labels": [i[0] for i in incident_data],
        "values": [i[1] for i in incident_data]
    }

    return {
        "revenue_chart": revenue_chart,
        "status_chart": status_chart,
        "incident_chart": incident_chart
    }

def get_company_stats(db: Session, company_id: int) -> Dict[str, Any]:
    total_requests = db.query(func.count(RescueRequest.id)).filter(RescueRequest.company_id == company_id).scalar()
    completed_requests = db.query(func.count(RescueRequest.id)).filter(
        RescueRequest.company_id == company_id, 
        RescueRequest.status == "COMPLETED"
    ).scalar()
    
    # Simple revenue calculation for company
    company_revenue = db.query(func.sum(Payment.amount)).join(RescueRequest).filter(
        RescueRequest.company_id == company_id,
        Payment.status == "success"
    ).scalar() or 0
    
    return {
        "total_requests": total_requests,
        "completed_requests": completed_requests,
        "revenue": float(company_revenue)
    }

def get_requests_for_export(db: Session):
    from app.models.user import User
    from app.models.company import RescueCompany
    
    results = db.query(
        RescueRequest.id,
        RescueRequest.created_at,
        User.full_name.label("customer"),
        RescueCompany.company_name.label("company"),
        RescueRequest.incident_type,
        RescueRequest.status,
        RescueRequest.agreed_price
    ).join(User, RescueRequest.user_id == User.id)\
     .outerjoin(RescueCompany, RescueRequest.company_id == RescueCompany.id)\
     .order_by(RescueRequest.created_at.desc()).all()
    
    return [
        {
            "ID": r.id,
            "Ngày tạo": r.created_at.strftime("%Y-%m-%d %H:%M"),
            "Khách hàng": r.customer,
            "Công ty": r.company or "N/A",
            "Loại sự cố": r.incident_type,
            "Trạng thái": r.status,
            "Chi phí": r.agreed_price or 0
        } for r in results
    ]


def get_request_report(
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    company_id: Optional[int] = None,
    incident_type: Optional[str] = None,
    status: Optional[str] = None,
) -> Dict[str, Any]:
    base = db.query(RescueRequest)
    base = _apply_request_filters(base, from_date, to_date, company_id, incident_type, status)

    total_requests = base.count()

    by_status = {k: 0 for k in _STATUS_BUCKETS}
    status_rows = (
        db.query(RescueRequest.status, func.count(RescueRequest.id))
        .filter(RescueRequest.id.in_(base.with_entities(RescueRequest.id)))
        .group_by(RescueRequest.status)
        .all()
    )
    for raw_status, count in status_rows:
        bucket = _map_request_status(raw_status)
        if bucket in by_status:
            by_status[bucket] += count
        else:
            by_status[bucket] = by_status.get(bucket, 0) + count

    incident_rows = (
        db.query(RescueRequest.incident_type, func.count(RescueRequest.id))
        .filter(RescueRequest.id.in_(base.with_entities(RescueRequest.id)))
        .group_by(RescueRequest.incident_type)
        .all()
    )
    by_incident_type = [{"type": t or "Khác", "count": c} for t, c in incident_rows]

    date_rows = (
        db.query(
            func.strftime("%Y-%m-%d", RescueRequest.created_at).label("date"),
            func.count(RescueRequest.id),
        )
        .filter(RescueRequest.id.in_(base.with_entities(RescueRequest.id)))
        .group_by("date")
        .order_by("date")
        .all()
    )
    by_date = [{"date": d, "count": c} for d, c in date_rows]

    cancelled = by_status.get("CANCELLED", 0)
    rejected = by_status.get("REJECTED", 0)
    cancel_rate = round((cancelled / total_requests) * 100, 2) if total_requests else 0.0
    reject_rate = round((rejected / total_requests) * 100, 2) if total_requests else 0.0

    return {
        "total_requests": total_requests,
        "by_status": by_status,
        "by_incident_type": by_incident_type,
        "by_date": by_date,
        "cancel_rate": cancel_rate,
        "reject_rate": reject_rate,
    }


def get_revenue_report(
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
) -> Dict[str, Any]:
    paid = db.query(Payment).filter(Payment.status == "success")
    paid = _apply_payment_date_filters(paid, from_date, to_date)

    total_revenue = paid.with_entities(func.sum(Payment.amount)).scalar() or 0

    company_rows = (
        db.query(
            RescueCompany.company_name,
            func.sum(Payment.amount).label("revenue"),
            func.count(func.distinct(RescueRequest.id)).label("request_count"),
        )
        .join(RescueRequest, RescueRequest.company_id == RescueCompany.id)
        .join(Payment, Payment.rescue_request_id == RescueRequest.id)
        .filter(Payment.status == "success")
    )
    if from_date:
        company_rows = company_rows.filter(Payment.created_at >= from_date)
    if to_date:
        company_rows = company_rows.filter(Payment.created_at <= to_date)
    company_rows = (
        company_rows.group_by(RescueCompany.id)
        .order_by(func.sum(Payment.amount).desc())
        .all()
    )
    by_company = [
        {
            "company_name": r.company_name,
            "revenue": float(r.revenue or 0),
            "request_count": r.request_count,
        }
        for r in company_rows
    ]

    date_rows = (
        db.query(
            func.strftime("%Y-%m-%d", Payment.created_at).label("date"),
            func.sum(Payment.amount),
        )
        .filter(Payment.status == "success")
    )
    date_rows = _apply_payment_date_filters(date_rows, from_date, to_date)
    date_rows = date_rows.group_by("date").order_by("date").all()
    by_date = [{"date": d, "revenue": float(v or 0)} for d, v in date_rows]

    method_rows = (
        db.query(
            Payment.payment_method,
            func.count(Payment.id),
            func.sum(Payment.amount),
        )
        .filter(Payment.status == "success")
    )
    method_rows = _apply_payment_date_filters(method_rows, from_date, to_date)
    method_rows = method_rows.group_by(Payment.payment_method).all()
    by_payment_method = [
        {
            "method": m or "unknown",
            "count": c,
            "total_amount": float(a or 0),
        }
        for m, c, a in method_rows
    ]

    return {
        "total_revenue": float(total_revenue),
        "by_company": by_company,
        "by_date": by_date,
        "by_payment_method": by_payment_method,
    }


def get_satisfaction_report(
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
) -> Dict[str, Any]:
    review_q = db.query(Review)
    review_q = _apply_review_date_filters(review_q, from_date, to_date)

    system_avg_rating = (
        review_q.with_entities(func.avg(Review.rating)).scalar() or 0.0
    )

    star_rows = (
        db.query(Review.rating, func.count(Review.id))
        .filter(Review.id.in_(review_q.with_entities(Review.id)))
        .group_by(Review.rating)
        .all()
    )
    by_star = {str(i): 0 for i in range(1, 6)}
    for rating, count in star_rows:
        by_star[str(rating)] = count

    def _company_ranking(order_desc: bool):
        q = db.query(RescueCompany).filter(RescueCompany.rating_count > 0)
        if order_desc:
            q = q.order_by(RescueCompany.rating_avg.desc(), RescueCompany.rating_count.desc())
        else:
            q = q.order_by(RescueCompany.rating_avg.asc(), RescueCompany.rating_count.desc())
        return [
            {
                "company_name": c.company_name,
                "rating_avg": float(c.rating_avg or 0),
                "rating_count": c.rating_count or 0,
            }
            for c in q.limit(5).all()
        ]

    top5_highest = _company_ranking(True)
    top5_lowest = _company_ranking(False)

    date_rows = (
        db.query(
            func.strftime("%Y-%m-%d", Review.created_at).label("date"),
            func.count(Review.id),
        )
        .filter(Review.id.in_(review_q.with_entities(Review.id)))
        .group_by("date")
        .order_by("date")
        .all()
    )
    reviews_by_date = [{"date": d, "count": c} for d, c in date_rows]

    return {
        "system_avg_rating": round(float(system_avg_rating), 2),
        "by_star": by_star,
        "top5_highest": top5_highest,
        "top5_lowest": top5_lowest,
        "reviews_by_date": reviews_by_date,
    }


def _report_title(report_type: str) -> str:
    return {
        "requests": "Báo cáo yêu cầu cứu hộ",
        "revenue": "Báo cáo doanh thu",
        "satisfaction": "Báo cáo mức độ hài lòng",
    }.get(report_type, "Báo cáo hệ thống")


def build_excel_report(
    db: Session,
    report_type: str,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    company_id: Optional[int] = None,
    incident_type: Optional[str] = None,
    status: Optional[str] = None,
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Tổng quan"

    period = f"{from_date.strftime('%Y-%m-%d') if from_date else '—'} → {to_date.strftime('%Y-%m-%d') if to_date else '—'}"
    ws_overview.append(["Báo cáo", _report_title(report_type)])
    ws_overview.append(["Khoảng thời gian", period])

    ws_company = wb.create_sheet("Theo công ty")
    ws_daily = wb.create_sheet("Theo ngày")

    if report_type == "requests":
        data = get_request_report(db, from_date, to_date, company_id, incident_type, status)
        ws_overview.append([])
        ws_overview.append(["Tổng yêu cầu", data["total_requests"]])
        ws_overview.append(["Tỉ lệ hủy (%)", data["cancel_rate"]])
        ws_overview.append(["Tỉ lệ từ chối (%)", data["reject_rate"]])
        for st, cnt in data["by_status"].items():
            ws_overview.append([f"Trạng thái {st}", cnt])
        ws_company.append(["Loại sự cố", "Số lượng"])
        for row in data["by_incident_type"]:
            ws_company.append([row["type"], row["count"]])
        ws_daily.append(["Ngày", "Số yêu cầu"])
        for row in data["by_date"]:
            ws_daily.append([row["date"], row["count"]])
    elif report_type == "revenue":
        data = get_revenue_report(db, from_date, to_date)
        ws_overview.append([])
        ws_overview.append(["Tổng doanh thu", data["total_revenue"]])
        ws_company.append(["Công ty", "Doanh thu", "Số yêu cầu"])
        for row in data["by_company"]:
            ws_company.append([row["company_name"], row["revenue"], row["request_count"]])
        ws_daily.append(["Ngày", "Doanh thu"])
        for row in data["by_date"]:
            ws_daily.append([row["date"], row["revenue"]])
    else:
        data = get_satisfaction_report(db, from_date, to_date)
        ws_overview.append([])
        ws_overview.append(["Rating trung bình hệ thống", data["system_avg_rating"]])
        for star, cnt in data["by_star"].items():
            ws_overview.append([f"{star} sao", cnt])
        ws_company.append(["Công ty", "Rating TB", "Số đánh giá"])
        ws_company.append(["— Top cao nhất —", "", ""])
        for row in data["top5_highest"]:
            ws_company.append([row["company_name"], row["rating_avg"], row["rating_count"]])
        ws_company.append([])
        ws_company.append(["— Top thấp nhất —", "", ""])
        for row in data["top5_lowest"]:
            ws_company.append([row["company_name"], row["rating_avg"], row["rating_count"]])
        ws_daily.append(["Ngày", "Số đánh giá"])
        for row in data["reviews_by_date"]:
            ws_daily.append([row["date"], row["count"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report(
    db: Session,
    report_type: str,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    company_id: Optional[int] = None,
    incident_type: Optional[str] = None,
    status: Optional[str] = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story: List = []

    period = f"{from_date.strftime('%Y-%m-%d') if from_date else '—'} → {to_date.strftime('%Y-%m-%d') if to_date else '—'}"
    story.append(Paragraph("ITSS — Hệ thống cứu hộ giao thông", styles["Title"]))
    story.append(Paragraph(_report_title(report_type), styles["Heading2"]))
    story.append(Paragraph(f"Khoảng thời gian: {period}", styles["Normal"]))
    story.append(Spacer(1, 12))

    if report_type == "requests":
        data = get_request_report(db, from_date, to_date, company_id, incident_type, status)
        rows = [
            ["Chỉ số", "Giá trị"],
            ["Tổng yêu cầu", str(data["total_requests"])],
            ["Tỉ lệ hủy (%)", str(data["cancel_rate"])],
            ["Tỉ lệ từ chối (%)", str(data["reject_rate"])],
        ]
        for st, cnt in data["by_status"].items():
            rows.append([f"Trạng thái {st}", str(cnt)])
    elif report_type == "revenue":
        data = get_revenue_report(db, from_date, to_date)
        rows = [["Chỉ số", "Giá trị"], ["Tổng doanh thu", f"{data['total_revenue']:,.0f} đ"]]
        rows.append(["", ""])
        rows.append(["Công ty", "Doanh thu", "Số yêu cầu"])
        for row in data["by_company"][:15]:
            rows.append([
                row["company_name"],
                f"{row['revenue']:,.0f}",
                str(row["request_count"]),
            ])
    else:
        data = get_satisfaction_report(db, from_date, to_date)
        rows = [
            ["Chỉ số", "Giá trị"],
            ["Rating TB hệ thống", str(data["system_avg_rating"])],
        ]
        for star, cnt in data["by_star"].items():
            rows.append([f"{star} sao", str(cnt)])

    table = Table(rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buf.getvalue()
